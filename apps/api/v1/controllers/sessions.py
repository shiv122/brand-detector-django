"""
Sessions Controller - List and inspect stored processing runs
"""

import os
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from django.conf import settings
from django.http import FileResponse
from django.urls import re_path
from rest_framework.decorators import api_view
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema, OpenApiParameter
from drf_spectacular.types import OpenApiTypes
from django.db.models import Count, Avg, Prefetch, Q

from apps.core.models import ProcessingSession, Frame, Detection, Classification
from apps.api.v1.controllers.ocr import _resolve_prompt
from apps.api.v1.shared_services import _ocr_service
from apps.services.ocr.ocr_queue import enqueue_ocr_job


def optional_slash_path(route, view, name=None):
    return [re_path(rf"^{route}/?$", view, name=name)]


def _session_type(session: ProcessingSession) -> str:
    return "video" if session.video_path else "image"


def _serialize_classifications(detection: Detection):
    return [
        {
            "class_id": c.class_id,
            "class_name": c.class_name,
            "confidence": c.confidence,
            "rank": c.rank,
        }
        for c in detection.classifications.all()
    ]


def _serialize_detection(detection: Detection):
    return {
        "id": detection.id,
        "class_id": detection.class_id,
        "class_name": detection.class_name,
        "confidence": detection.confidence,
        "bbox": [
            detection.bbox_x1,
            detection.bbox_y1,
            detection.bbox_x2,
            detection.bbox_y2,
        ],
        "classification": _serialize_classifications(detection),
    }


def _serialize_frame(frame: Frame):
    return {
        "id": frame.id,
        "frame_number": frame.frame_number,
        "frame_url": frame.frame_url,
        "timestamp": frame.timestamp,
        "total_detections": frame.total_detections,
        "ocr_summary": frame.ocr_summary,
        "detections": [_serialize_detection(d) for d in frame.detections.all()],
    }


@extend_schema(
    summary="List processing sessions",
    description="Paginated list of all processing sessions (runs) with aggregate stats.",
    parameters=[
        OpenApiParameter(name="status", type=OpenApiTypes.STR, location=OpenApiParameter.QUERY, required=False),
        OpenApiParameter(name="type", type=OpenApiTypes.STR, location=OpenApiParameter.QUERY, required=False, description="video | image"),
        OpenApiParameter(name="search", type=OpenApiTypes.STR, location=OpenApiParameter.QUERY, required=False),
        OpenApiParameter(name="limit", type=OpenApiTypes.INT, location=OpenApiParameter.QUERY, required=False),
        OpenApiParameter(name="offset", type=OpenApiTypes.INT, location=OpenApiParameter.QUERY, required=False),
    ],
)
@api_view(["GET"])
def list_sessions(request):
    qs = ProcessingSession.objects.all()

    status_filter = request.GET.get("status")
    if status_filter:
        qs = qs.filter(status=status_filter)

    type_filter = request.GET.get("type")
    if type_filter == "video":
        qs = qs.filter(video_path__isnull=False).exclude(video_path="")
    elif type_filter == "image":
        qs = qs.filter(video_path__isnull=True)

    search = (request.GET.get("search") or "").strip()
    if search:
        qs = qs.filter(video_filename__icontains=search) | qs.filter(session_id__icontains=search)

    qs = qs.order_by("-created_at")

    total = qs.count()  # plain COUNT over sessions — no joins

    try:
        limit = max(1, min(int(request.GET.get("limit", 50)), 200))
    except (TypeError, ValueError):
        limit = 50
    try:
        offset = max(0, int(request.GET.get("offset", 0)))
    except (TypeError, ValueError):
        offset = 0

    page = list(qs[offset : offset + limit])
    ids = [s.id for s in page]

    # Aggregate detections and frames in TWO separate grouped queries — NOT one
    # multi-annotate. Annotating Count(detections) AND Count(frames) on the same
    # queryset joins detections x frames per session (a cartesian explosion that
    # makes this endpoint crawl on full-length videos). Two grouped queries over
    # only the current page's sessions stay flat and index-friendly.
    det_stats = {
        r["session"]: r
        for r in Detection.objects.filter(session_id__in=ids)
        .values("session")
        .annotate(c=Count("id"), uc=Count("class_name", distinct=True), ac=Avg("confidence"))
    }
    frame_stats = {
        r["session"]: r
        for r in Frame.objects.filter(session_id__in=ids)
        .values("session")
        .annotate(fc=Count("id"), oc=Count("id", filter=Q(ocr_summary__isnull=False)))
    }

    sessions = []
    for s in page:
        d = det_stats.get(s.id) or {}
        f = frame_stats.get(s.id) or {}
        progress = 0
        if s.total_frames > 0:
            progress = int((s.processed_frames / s.total_frames) * 100)
        sessions.append(
            {
                "id": s.id,
                "session_id": s.session_id,
                "type": _session_type(s),
                "name": s.video_filename or f"session_{s.session_id[:8]}",
                "status": s.status,
                "total_frames": s.total_frames,
                "processed_frames": s.processed_frames,
                "progress": progress,
                "frames_stored": f.get("fc", 0),
                "total_detections": d.get("c", 0),
                "unique_brands": d.get("uc", 0),
                "avg_confidence": d.get("ac"),
                "frames_per_second": s.frames_per_second,
                "confidence_threshold": s.confidence_threshold,
                "created_at": s.created_at.isoformat(),
                "completed_at": s.completed_at.isoformat() if s.completed_at else None,
                "has_processed_video": bool(s.processed_video_path),
                # OCR: whether it was enabled for the run, and how many frames
                # have an OCR result so far (so the UI can show "OCR n/N").
                "ocr_enabled": bool(
                    (s.run_params or {}).get("enable_ocr")
                    or (s.settings or {}).get("enable_ocr")
                ),
                "ocr_frames": f.get("oc", 0),
            }
        )

    return Response(
        {
            "sessions": sessions,
            "total": total,
            "limit": limit,
            "offset": offset,
        }
    )


@extend_schema(
    summary="Get full session detail",
    description="Returns the session metadata plus all frames, detections, and classifications.",
    parameters=[
        OpenApiParameter(name="session_id", type=OpenApiTypes.STR, location=OpenApiParameter.PATH),
    ],
)
@api_view(["GET"])
def session_detail(request, session_id):
    try:
        session = ProcessingSession.objects.get(session_id=session_id)
    except ProcessingSession.DoesNotExist:
        return Response({"error": "Session not found"}, status=404)

    frames_qs = (
        session.frames.all()
        .order_by("frame_number")
        .prefetch_related(
            Prefetch(
                "detections",
                queryset=Detection.objects.order_by("-confidence").prefetch_related(
                    Prefetch(
                        "classifications",
                        queryset=Classification.objects.order_by("rank"),
                    )
                ),
            )
        )
    )

    frames = [_serialize_frame(f) for f in frames_qs]

    detections_qs = session.detections.all()
    logo_totals = {}
    for row in detections_qs.values("class_name").annotate(count=Count("id")):
        logo_totals[row["class_name"]] = row["count"]

    # RQ job ids for frames whose OCR still needs to land (no result yet, or an
    # error result) AND that were actually enqueued. The detail page polls these
    # via /ocr/jobs so that, after a resume, OCR shows as "running" (and the
    # Resume button stays hidden) even on a fresh page load — and self-heals to
    # "stalled" again if those jobs turn out to be gone/expired.
    needs_ocr_q = Q(ocr_summary__isnull=True) | Q(ocr_summary__has_key="error")
    ocr_pending_job_ids = list(
        session.frames.filter(needs_ocr_q)
        .exclude(ocr_job_id="")
        .order_by("frame_number")
        .values_list("ocr_job_id", flat=True)[:5000]
    )

    return Response(
        {
            "session": {
                "id": session.id,
                "session_id": session.session_id,
                "type": _session_type(session),
                "name": session.video_filename or f"session_{session.session_id[:8]}",
                "video_filename": session.video_filename,
                "status": session.status,
                "total_frames": session.total_frames,
                "processed_frames": session.processed_frames,
                "frames_per_second": session.frames_per_second,
                "confidence_threshold": session.confidence_threshold,
                "settings": session.settings,
                "processed_video_url": (
                    f"/static/{os.path.basename(session.processed_video_path)}"
                    if session.processed_video_path
                    else None
                ),
                "created_at": session.created_at.isoformat(),
                "updated_at": session.updated_at.isoformat(),
                "completed_at": (
                    session.completed_at.isoformat() if session.completed_at else None
                ),
                # Whether OCR was configured for this run, so the UI can offer to
                # resume it even when zero frames have a result yet.
                "ocr_enabled": bool(
                    (session.run_params or {}).get("enable_ocr")
                    or (session.settings or {}).get("enable_ocr")
                ),
            },
            "frames": frames,
            "logo_totals": logo_totals,
            "total_detections": detections_qs.count(),
            "unique_logos": list(logo_totals.keys()),
            "ocr_pending_job_ids": ocr_pending_job_ids,
        }
    )


@extend_schema(
    summary="Resume / run OCR for a session's frames",
    description=(
        "Enqueues OCR for every stored frame in this session that has no "
        "successful OCR result yet — i.e. frames whose ocr_summary is missing "
        "or holds an error. This resumes from the last frame where OCR "
        "succeeded; if OCR never ran, every frame is queued. The run's "
        "original OCR prompt (stored in run_params) is reused. Pass "
        "`prompt` / `prompt_slug` / `sport` to override it, or to supply one "
        "when the run had no OCR configured. Results are written back to "
        "Frame.ocr_summary and are pollable via /ocr/jobs."
    ),
    parameters=[
        OpenApiParameter(name="session_id", type=OpenApiTypes.STR, location=OpenApiParameter.PATH),
    ],
)
@api_view(["POST"])
def resume_session_ocr(request, session_id):
    try:
        session = ProcessingSession.objects.get(session_id=session_id)
    except ProcessingSession.DoesNotExist:
        return Response({"error": "Session not found"}, status=404)

    if not _ocr_service.is_available():
        return Response(
            {"error": "GLM OCR endpoint is not configured — set GLM_OCR_HOST."},
            status=503,
        )

    data = request.data or {}
    params = session.run_params or {}

    # Prompt resolution: an explicit override in the body wins; otherwise reuse
    # the prompt the run was started with (persisted in run_params). A run that
    # never had OCR configured has no stored prompt, so the caller must supply
    # one.
    override = any(
        (data.get("prompt"), data.get("prompt_slug"), data.get("sport"))
    )
    if override:
        prompt, prompt_meta, err = _resolve_prompt(
            data.get("prompt"), data.get("prompt_slug"), data.get("sport")
        )
        if err is not None:
            return err
    else:
        prompt = params.get("ocr_prompt")
        prompt_meta = params.get("ocr_meta") or {}
        if not prompt:
            return Response(
                {
                    "error": "this run has no stored OCR prompt — pass "
                    "prompt / prompt_slug / sport to run OCR."
                },
                status=422,
            )

    # Frames that still need OCR: no result yet, or a previous error result.
    # Ordering by frame_number isn't required for correctness (jobs run
    # concurrently) but keeps the enqueue order intuitive.
    needs = (
        session.frames.all()
        .filter(Q(ocr_summary__isnull=True) | Q(ocr_summary__has_key="error"))
        .order_by("frame_number")
    )

    enqueued = 0
    skipped_no_url = 0
    job_ids: list[str] = []
    to_update: list[Frame] = []
    for frame in needs.iterator():
        url = frame.frame_url or ""
        # GLM OCR fetches by URL, so a frame stored only on local disk can't be
        # OCR'd remotely — skip it rather than enqueue a job that will fail.
        if not url.startswith(("http://", "https://")):
            skipped_no_url += 1
            continue
        job = enqueue_ocr_job(
            image_url=url, prompt=prompt, prompt_meta=prompt_meta, frame_id=frame.id
        )
        jid = (job or {}).get("id")
        if jid:
            job_ids.append(jid)
            frame.ocr_job_id = jid
            to_update.append(frame)
            enqueued += 1

    if to_update:
        Frame.objects.bulk_update(to_update, ["ocr_job_id"], batch_size=500)

    # Persist the prompt + enable flag so subsequent resumes work and the
    # sessions list/badge reflects OCR being active for this run.
    if enqueued and (override or not params.get("enable_ocr")):
        params["enable_ocr"] = True
        params["ocr_prompt"] = prompt
        params["ocr_meta"] = prompt_meta
        session.run_params = params
        session.save(update_fields=["run_params", "updated_at"])

    return Response(
        {
            "session_id": session.session_id,
            "enqueued": enqueued,
            "skipped_no_url": skipped_no_url,
            "ocr_job_ids": job_ids,
            "prompt_meta": prompt_meta,
        }
    )


def _ocr_raw_text(ocr_summary) -> str:
    """Extract a flat string of OCR raw text from the stored ocr_summary JSON."""
    if not isinstance(ocr_summary, dict):
        return ""
    raw = ocr_summary.get("raw_text")
    if isinstance(raw, str):
        return raw.strip()
    if isinstance(raw, list):
        return " ".join(str(x) for x in raw).strip()
    return ""


def _xlsx_export_path(session: ProcessingSession) -> Path:
    static_dir = getattr(settings, "STATIC_ROOT", None) or os.path.join(
        settings.BASE_DIR, "staticfiles"
    )
    out_dir = Path(static_dir) / "csv_reports"
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_id = session.session_id[:8]
    return out_dir / f"detection_report_{timestamp}_{safe_id}.xlsx"


@extend_schema(
    summary="Export session as XLSX (project report format)",
    description=(
        "Generate an Excel workbook with one row per (frame, brand) aggregation. "
        "Columns: Project Name, Image, Brand, Instances, Size, OCR Raw Text."
    ),
    parameters=[
        OpenApiParameter(name="session_id", type=OpenApiTypes.STR, location=OpenApiParameter.PATH),
    ],
)
@api_view(["POST"])
def export_session_xlsx(request, session_id):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return Response({"error": "openpyxl not available on the server"}, status=500)

    try:
        session = ProcessingSession.objects.get(session_id=session_id)
    except ProcessingSession.DoesNotExist:
        return Response({"error": "Session not found"}, status=404)

    project_name = session.video_filename or f"session_{session.session_id[:8]}"

    frames_qs = (
        session.frames.all()
        .order_by("frame_number")
        .prefetch_related(
            Prefetch(
                "detections",
                queryset=Detection.objects.prefetch_related("classifications"),
            )
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Detections"
    headers = [
        "Project Name",
        "Frame #",
        "Timestamp (s)",
        "Image",
        "Brand",
        "Instances",
        "Classification",
        "Size",
        "Raw Text",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D2D2D")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    row_idx = 2
    for frame in frames_qs:
        detections = list(frame.detections.all())
        if not detections:
            continue

        # Image name: local path when present, else the Spaces URL's filename
        # (the Spaces flow stores frame_path="" and keeps the URL on frame_url).
        image_name = os.path.basename((frame.frame_path or frame.frame_url or "").split("?")[0])
        ocr_text = _ocr_raw_text(frame.ocr_summary)

        # Group detections by brand within this frame
        by_brand: dict[str, list[Detection]] = defaultdict(list)
        for det in detections:
            by_brand[det.class_name].append(det)

        for brand, dets in by_brand.items():
            instances = len(dets)
            # Representative detection for size — largest-area instance.
            rep = max(
                dets,
                key=lambda d: (d.bbox_x2 - d.bbox_x1) * (d.bbox_y2 - d.bbox_y1),
            )
            bw = int(round(rep.bbox_x2 - rep.bbox_x1))
            bh = int(round(rep.bbox_y2 - rep.bbox_y1))
            size_str = f"{bw}X{bh}"

            # Classification: top-1 asset class per instance, counted across the
            # group, e.g. "perimeter_board x3, caddie_bib x1".
            cls_counts: dict[str, int] = defaultdict(int)
            for det in dets:
                ranked = list(det.classifications.all())  # prefetched, rank order
                if ranked:
                    cls_counts[ranked[0].class_name] += 1
            classification_str = (
                ", ".join(
                    f"{name} x{cnt}"
                    for name, cnt in sorted(cls_counts.items(), key=lambda kv: -kv[1])
                )
                or "N/A"
            )

            ws.cell(row=row_idx, column=1, value=project_name)
            ws.cell(row=row_idx, column=2, value=frame.frame_number)
            ws.cell(row=row_idx, column=3, value=round(frame.timestamp or 0, 2))
            ws.cell(row=row_idx, column=4, value=image_name)
            ws.cell(row=row_idx, column=5, value=brand)
            ws.cell(row=row_idx, column=6, value=instances)
            ws.cell(row=row_idx, column=7, value=classification_str)
            ws.cell(row=row_idx, column=8, value=size_str)
            ws.cell(row=row_idx, column=9, value=ocr_text)
            row_idx += 1

    # Column widths — keep it readable; classification + OCR get more space.
    widths = [22, 8, 12, 26, 22, 10, 30, 12, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"

    out_path = _xlsx_export_path(session)
    wb.save(out_path)

    return Response(
        {
            "message": "XLSX exported successfully",
            "session_id": session.session_id,
            "filename": out_path.name,
            "url": f"/static/csv_reports/{out_path.name}",
            "rows": row_idx - 2,
        }
    )


@extend_schema(
    summary="Download an exported XLSX file",
    parameters=[
        OpenApiParameter(name="filename", type=OpenApiTypes.STR, location=OpenApiParameter.PATH),
    ],
)
@api_view(["GET"])
def download_xlsx(request, filename):
    static_dir = getattr(settings, "STATIC_ROOT", None) or os.path.join(
        settings.BASE_DIR, "staticfiles"
    )
    file_path = Path(static_dir) / "csv_reports" / filename
    if not file_path.exists() or not file_path.suffix.lower() == ".xlsx":
        return Response({"error": "File not found"}, status=404)
    return FileResponse(
        open(file_path, "rb"),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        as_attachment=True,
    )


urlpatterns = [
    *optional_slash_path("sessions", list_sessions, name="sessions-list"),
    *optional_slash_path(
        r"session/(?P<session_id>[^/]+)/detail",
        session_detail,
        name="session-detail",
    ),
    *optional_slash_path(
        r"session/(?P<session_id>[^/]+)/export-xlsx",
        export_session_xlsx,
        name="session-export-xlsx",
    ),
    *optional_slash_path(
        r"session/(?P<session_id>[^/]+)/resume-ocr",
        resume_session_ocr,
        name="session-resume-ocr",
    ),
    *optional_slash_path(
        r"xlsx-files/download/(?P<filename>[^/]+)",
        download_xlsx,
        name="xlsx-download",
    ),
]
